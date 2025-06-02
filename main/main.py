import cv2
import dlib
import os
import time
import numpy as np
from deepface import DeepFace
from pathlib import Path
import csv
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Thư viện cho Google Sheets
import gspread
from google.oauth2.service_account import Credentials # Nếu dùng Service Account
from google_auth_oauthlib.flow import InstalledAppFlow # Nếu dùng Desktop App OAuth

# --- CONFIG ---
SAVE_DIR = os.path.join(os.getcwd(), "Face_detection", "outputs")
VECTOR_DIR = os.path.join(os.getcwd(), "Feature_extractor", "output_vector")
REF_IMG_DIR = "../reference_images"
REF_VEC_DIR = "../reference_vectors"
MODEL_NAME = "VGG-Face"
DETECTOR_BACKEND = "opencv"
THRESHOLD = 0.6

# --- Cấu hình lưu trữ dữ liệu ---
ATTENDANCE_CSV_FILE = os.path.join(os.getcwd(), "attendance.csv")
ATTENDANCE_EXCEL_FILE = os.path.join(os.getcwd(), "attendance.xlsx")

# Google Sheets Config (cần cấu hình nếu muốn dùng)
# Đổi tên file credentials.json đã tải về
GOOGLE_SHEETS_CREDENTIALS_FILE = 'credentials.json'
SPREADSHEET_NAME = 'Attendance_Log' # Tên bảng tính Google Sheets của bạn
WORKSHEET_NAME = 'Điểm danh' # Tên sheet trong bảng tính

# --- SETUP ---
os.makedirs(SAVE_DIR, exist_ok=True)
os.makedirs(VECTOR_DIR, exist_ok=True)
os.makedirs(REF_VEC_DIR, exist_ok=True)

# --- FUNC: Cosine Distance ---
def find_cosine_distance(source_representation, test_representation):
    a = np.dot(source_representation, test_representation)
    b = np.sum(source_representation ** 2)
    c = np.sum(test_representation ** 2)
    return 1 - (a / (np.sqrt(b) * np.sqrt(c)))

# --- FUNC: Feature Extraction ---
def extract_feature(image_path):
    try:
        embedding_objs = DeepFace.represent(
            img_path=image_path,
            model_name=MODEL_NAME,
            detector_backend=DETECTOR_BACKEND,
            enforce_detection=False,
            align=True
        )
        if embedding_objs and len(embedding_objs) > 0:
            return np.array(embedding_objs[0]['embedding'])
        else:
            print(f"[!] Không trích xuất được đặc trưng từ {image_path}")
            return None
    except Exception as e:
        print(f"[!] Lỗi khi trích xuất đặc trưng: {e}")
        return None

# --- FUNC: Chuẩn bị vector mẫu (1 lần hoặc nếu chưa có) ---
def prepare_reference_vectors():
    # Đảm bảo thư mục reference_images tồn tại
    if not os.path.exists(REF_IMG_DIR):
        print(f"[ERROR] Thư mục ảnh mẫu không tồn tại: {REF_IMG_DIR}")
        print("Vui lòng tạo thư mục này và đặt ảnh mẫu vào.")
        return

    for filename in os.listdir(REF_IMG_DIR):
        if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
            name = os.path.splitext(filename)[0]
            vec_path = os.path.join(REF_VEC_DIR, f"{name}.npy")
            if not os.path.exists(vec_path):
                img_path = os.path.join(REF_IMG_DIR, filename)
                vec = extract_feature(img_path)
                if vec is not None:
                    np.save(vec_path, vec)
                    print(f"[✓] Đã lưu vector mẫu: {vec_path}")
                else:
                    print(f"[!] Không thể tạo vector cho ảnh {filename}")
            else:
                print(f"[INFO] Vector cho {name} đã tồn tại, bỏ qua.")


# --- FUNC: So sánh với tất cả vector mẫu ---
def compare_with_all_references(test_vector):
    # Đảm bảo thư mục reference_vectors tồn tại
    if not os.path.exists(REF_VEC_DIR) or not os.listdir(REF_VEC_DIR):
        print("[WARNING] Thư mục vector mẫu rỗng hoặc không tồn tại. Vui lòng chạy prepare_reference_vectors trước.")
        return None

    for vec_file in os.listdir(REF_VEC_DIR):
        if vec_file.endswith(".npy"):
            ref_vector = np.load(os.path.join(REF_VEC_DIR, vec_file))
            dist = find_cosine_distance(ref_vector, test_vector)
            if dist <= THRESHOLD:
                name = os.path.splitext(vec_file)[0]
                print(f"[=] Khớp với {name} (distance={dist:.4f})")
                return name
    return None

# --- NEW FUNC: Lưu dữ liệu vào CSV ---
def save_to_csv(data_entry, filename=ATTENDANCE_CSV_FILE):
    headers = ["ID_NhanVien", "Ten_NhanVien", "ThoiGian", "TrangThai", "GhiChu"]
    file_exists = os.path.isfile(filename)

    with open(filename, 'a', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        if not file_exists:
            writer.writeheader() # Ghi header nếu file chưa tồn tại
        writer.writerow(data_entry)
    print(f"[✓] Đã ghi vào CSV: {data_entry}")

# --- NEW FUNC: Lưu dữ liệu vào Excel ---
def save_to_excel(data_entry, filename=ATTENDANCE_EXCEL_FILE):
    headers = ["ID_NhanVien", "Ten_NhanVien", "ThoiGian", "TrangThai", "GhiChu"]
    try:
        # Tải workbook nếu tồn tại, nếu không thì tạo mới
        if os.path.exists(filename):
            workbook = load_workbook(filename)
            sheet = workbook.active
            # Kiểm tra và thêm header nếu sheet trống
            if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
                sheet.append(headers)
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Điểm danh"
            sheet.append(headers)

        # Thêm dữ liệu
        row_data = [data_entry.get(h, '') for h in headers]
        sheet.append(row_data)

        # Điều chỉnh độ rộng cột tự động (tùy chọn)
        for col_idx, col_name in enumerate(headers, 1):
            max_length = 0
            for row in sheet.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        workbook.save(filename)
        print(f"[✓] Đã ghi vào Excel: {data_entry}")
    except Exception as e:
        print(f"[ERROR] Lỗi khi ghi vào Excel: {e}")

# --- NEW FUNC: Lưu dữ liệu vào Google Sheets ---
def save_to_google_sheets(data_entry):
    headers = ["ID_NhanVien", "Ten_NhanVien", "ThoiGian", "TrangThai", "GhiChu"]
    try:
        # Xác thực với Google API
        # Nếu dùng Desktop App OAuth:
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        flow = InstalledAppFlow.from_client_secrets_file(GOOGLE_SHEETS_CREDENTIALS_FILE, scope)
        creds = flow.run_local_server(port=0)
        gc = gspread.authorize(creds)

        # Nếu dùng Service Account (phức tạp hơn cho người mới):
        # creds = Credentials.from_service_account_file(
        #     GOOGLE_SHEETS_CREDENTIALS_FILE,
        #     scopes=['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive'])
        # gc = gspread.authorize(creds)

        spreadsheet = gc.open(SPREADSHEET_NAME)
        try:
            worksheet = spreadsheet.worksheet(WORKSHEET_NAME)
        except gspread.exceptions.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(title=WORKSHEET_NAME, rows="100", cols="20")
            print(f"[INFO] Đã tạo sheet mới: {WORKSHEET_NAME}")

        # Ghi header nếu sheet trống hoặc không có header
        if not worksheet.row_values(1) or worksheet.row_values(1)[0] != headers[0]:
            worksheet.clear() # Xóa nội dung cũ nếu muốn ghi header mới
            worksheet.append_row(headers)

        # Thêm dữ liệu
        row_data = [data_entry.get(h, '') for h in headers]
        worksheet.append_row(row_data)
        print(f"[✓] Đã ghi vào Google Sheets: {data_entry}")

    except FileNotFoundError:
        print(f"[ERROR] File xác thực Google Sheets '{GOOGLE_SHEETS_CREDENTIALS_FILE}' không tìm thấy.")
        print("Vui lòng đảm bảo bạn đã tải file 'credentials.json' và đặt đúng chỗ.")
    except Exception as e:
        print(f"[ERROR] Lỗi khi ghi vào Google Sheets: {e}")


# --- MAIN LOOP ---
def main():
    prepare_reference_vectors()
    detector = dlib.get_frontal_face_detector()
    cap = cv2.VideoCapture(0)
    last_processed_person = {} # Lưu trữ thời gian cuối cùng của người được nhận diện
    processing_interval_seconds = 10 # Khoảng thời gian tối thiểu giữa các lần điểm danh của cùng một người

    if not cap.isOpened():
        print("[ERROR] Không mở được webcam! Có thể do webcam đang bận hoặc không có thiết bị webcam.")
        return # Thoát nếu không mở được webcam
    else:
        print("[INFO] Webcam mở thành công.")

    # Khởi tạo các file lưu trữ (nếu cần)
    # save_to_csv({}, init_only=True) # Không cần init_only vì hàm đã tự kiểm tra
    # save_to_excel({}, init_only=True) # Tương tự

    while True:
        ret, frame = cap.read()
        if not ret:
            print("[INFO] Mất kết nối webcam hoặc luồng video kết thúc.")
            break

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        faces = detector(rgb)
        current_time = time.time()

        for face in faces:
            x1, y1 = face.left(), face.top()
            x2, y2 = face.right(), face.bottom()
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)

            face_crop = rgb[y1:y2, x1:x2]
            # Đảm bảo ảnh crop không bị rỗng
            if face_crop.shape[0] == 0 or face_crop.shape[1] == 0:
                print("[WARNING] Ảnh crop khuôn mặt bị rỗng, bỏ qua.")
                continue

            # Để tránh ghi ảnh liên tục, chỉ lưu ảnh khi nhận diện thành công
            # timestamp = time.strftime("%Y%m%d_%H%M%S")
            # face_img_path = os.path.join(SAVE_DIR, f"face_{timestamp}.jpg")
            # cv2.imwrite(face_img_path, cv2.cvtColor(face_crop, cv2.COLOR_RGB2BGR)) # Comment out để tránh ghi quá nhiều ảnh

            # Trích xuất đặc trưng trực tiếp từ face_crop
            # (Bạn cần DeepFace để trích xuất từ numpy array. Nếu DeepFace không hỗ trợ trực tiếp,
            # bạn có thể cần lưu tạm ra file và đọc lại, nhưng việc này tốn thời gian.
            # Để đơn giản, tôi sẽ giữ cách lưu tạm file, hoặc dùng img_to_array)

            # Option 1: Lưu tạm file và đọc lại (cách hiện tại của bạn)
            temp_face_path = os.path.join(SAVE_DIR, "temp_face_crop.jpg")
            cv2.imwrite(temp_face_path, cv2.cvtColor(face_crop, cv2.COLOR_RGB2BGR))
            test_vector = extract_feature(temp_face_path)
            # os.remove(temp_face_path) # Xóa file tạm ngay sau khi sử dụng

            # Option 2: Nếu DeepFace hỗ trợ trích xuất từ numpy array (ít tốn I/O hơn)
            # test_vector = extract_feature_from_array(face_crop) # Giả định có hàm này

            label = "Unknown"
            if test_vector is not None:
                matched_name = compare_with_all_references(test_vector)
                if matched_name:
                    label = matched_name
                    # Kiểm tra xem đã xử lý người này trong khoảng thời gian gần đây chưa
                    if matched_name not in last_processed_person or \
                       (current_time - last_processed_person[matched_name]) > processing_interval_seconds:
                        
                        # Dữ liệu để lưu
                        id_nhan_vien = f"ID_{matched_name.replace(' ', '_')}" # Giả định ID từ tên file
                        thoi_gian = time.strftime("%Y-%m-%d %H:%M:%S")
                        data_entry = {
                            "ID_NhanVien": id_nhan_vien,
                            "Ten_NhanVien": matched_name,
                            "ThoiGian": thoi_gian,
                            "TrangThai": "Check-in",
                            "GhiChu": ""
                        }

                        # LƯU DỮ LIỆU VÀO CÁC ĐỊNH DẠNG:
                        save_to_csv(data_entry)
                        save_to_excel(data_entry)
                        # Uncomment dòng dưới để lưu vào Google Sheets
                        # save_to_google_sheets(data_entry)

                        last_processed_person[matched_name] = current_time # Cập nhật thời gian xử lý

            cv2.putText(frame, label, (x1, y1 - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8,
                        (0, 255, 0) if label != "Unknown" else (0, 0, 255), 2)

        cv2.imshow("Face Recognition", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    main()
